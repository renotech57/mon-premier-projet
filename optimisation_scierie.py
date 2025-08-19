
import sys, os
sys.path.append(r"C:\cadwork\libs")
from typing import Any, Tuple, List, Dict
from datetime import datetime
from collections import defaultdict

# === Imports Cadwork & libs tierces ===
import cadwork
import utility_controller as uc
import material_controller as mc
import attribute_controller as ac
import element_controller as ec
import geometry_controller as gc

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import subprocess
import json

# =========================================================
# Utils
# =========================================================

MARGE_COUPE_DEFAULT = 80

def safe_float(val, default=0.0):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).strip().replace('\\u00A0','').replace(' ','').replace(',','.')
        return float(s) if s else default
    except Exception:
        return default

def log_message(message, level="INFO"):
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {level}: {message}")

# conversions
def _mm_to_m(v_mm: Any) -> float:    return safe_float(v_mm) / 1000.0
def _mm2_to_m2(v_mm2: Any) -> float: return safe_float(v_mm2) / 1_000_000.0
def _mm3_to_m3(v_mm3: Any) -> float: return safe_float(v_mm3) / 1_000_000_000.0

def _cad_to_unit(val, unit):
    u = (unit or '').lower()
    if u == 'ml': return _mm_to_m(val)
    if u == 'm2': return _mm2_to_m2(val)
    if u == 'm3': return _mm3_to_m3(val)
    return safe_float(val)

# cotes de liste
def _get_list_length(eid: int) -> float:
    return safe_float(gc.get_list_length(eid) if hasattr(gc, 'get_list_length') else gc.get_length(eid))

def _get_phys_length(eid: int) -> float:
    # Longueur "physique" (avec dépassements/usages si Cadwork les inclut)
    return safe_float(gc.get_length(eid))

def _get_list_width(eid: int) -> float:
    return safe_float(gc.get_list_width(eid))

def _get_list_height(eid: int) -> float:
    return safe_float(gc.get_list_height(eid))

def _two_largest_face_dims_mm(eid: int) -> Tuple[float,float,float]:
    L = _get_list_length(eid)
    W = _get_list_width(eid)
    H = _get_list_height(eid)
    dims = sorted([L,W,H], reverse=True)
    return dims[0], dims[1], dims[2]  # (a_mm, b_mm, ep_mm)

# heuristique d’unité
def get_material_unit(material_name: str) -> str:
    try:
        mat_id = mc.get_material_id(material_name)
        if hasattr(mc, 'get_unit'):
            u = mc.get_unit(mat_id)
            if u: return u.lower()
    except Exception:
        pass

    up = (material_name or '').upper()
    if any(k in up for k in ['OSB','CONTREPLAQUE','PANNEAU','PLAQUE','ISOLANT']):
        return 'm2'
    if any(k in up for k in ['POUTRE','PROFILE','CHEVRON','TASSEAU','SJ-']):
        return 'ml'
    return 'm3'

def calculate_quantity_by_unit(largeur_mm, hauteur_mm, longueur_mm, unit):
    if unit == 'ml':
        return _mm_to_m(longueur_mm)
    if unit == 'm2':
        return _mm2_to_m2(safe_float(largeur_mm)*safe_float(hauteur_mm))
    return _mm3_to_m3(safe_float(largeur_mm)*safe_float(hauteur_mm)*safe_float(longueur_mm))

# =========================================================
# Calcul non optimisé
# =========================================================

def calculate_quantity_with_cadwork_method(eid, unit, method):
    """
    Non optimisé.
    m² : BRUT (2 plus grandes cotes) -> m²  [méthodes surface ignorées]
    m³ : respecte la méthode choisie
    ml : respecte la méthode choisie
    """
    try:
        u = (unit or '').lower()
        m = (method or 'manuel').lower()

        if u == 'm2':
            a_mm, b_mm, _ = _two_largest_face_dims_mm(eid)
            m2 = _mm2_to_m2(a_mm * b_mm)
            if m2 <= 0 and hasattr(gc, 'get_area_of_front_face'):
                return _mm2_to_m2(gc.get_area_of_front_face(eid))
            return m2

        if u == 'm3':
            if m == 'volume_physique_reel' and hasattr(gc, 'get_actual_physical_volume'):
                return _mm3_to_m3(gc.get_actual_physical_volume(eid))
            if m == 'volume_standard' and hasattr(gc, 'get_volume'):
                return _mm3_to_m3(gc.get_volume(eid))
            if m == 'volume_liste' and hasattr(gc, 'get_list_volume'):
                return _mm3_to_m3(gc.get_list_volume(eid))
            # Fallback manuel (enveloppe L×W×H)
            L = _get_list_length(eid); W = _get_list_width(eid); H = _get_list_height(eid)
            return _mm3_to_m3(L*W*H)

        if u == 'ml':
            # Respect de la méthode
            if m == 'longueur_liste' and hasattr(gc, 'get_list_length'):
                return _mm_to_m(gc.get_list_length(eid))
            if m == 'longueur_avec_depassement' and hasattr(gc, 'get_length'):
                return _mm_to_m(gc.get_length(eid))
            # Fallback "manuel" : on privilégie la longueur physique quand dispo
            if hasattr(gc, 'get_length'):
                return _mm_to_m(gc.get_length(eid))
            return _mm_to_m(_get_list_length(eid))

        # fallback
        return calculate_quantity_by_unit(_get_list_width(eid), _get_list_height(eid), _get_list_length(eid), unit)

    except Exception as e:
        log_message(f"Erreur calcul non optimisé eid={eid} unit={unit}: {e}", "WARNING")
        return calculate_quantity_by_unit(_get_list_width(eid), _get_list_height(eid), _get_list_length(eid), unit)

def is_materiau_13m(material_name: str) -> bool:
    materiaux_13m = ["KVH", "BMR", "SJ-60*39*200_L", "LVL", "SJ-60*39*240_L",
                     "SJ-60*39*300_L", "SJ-60*39*360_L", "SJ-60*39*400_L",
                     "SJ-60*39*450_L", "SJ-90*39*300_L", "SJ-90*39*360_L", "SJ-90*39*400_L"]
    up = (material_name or '').upper()
    return any(code in up for code in materiaux_13m)

# =========================================================
# Interface
# =========================================================

class OptimizationConfigUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🪓 Optimisation Scierie v5.2 — m² BRUT")
        self.root.geometry("1100x800")
        self.root.resizable(True, True)

        self.materiaux_configs: Dict[str,dict] = {}
        self.materiaux_detectes: List[str] = []
        self.current_material: str = None
        self.elements_data: List[dict] = []

        self.setup_variables()
        self.create_interface()
        self.detect_materials()
        self.load_default_configs()

    # ----------------------
    # Variables UI
    # ----------------------
    def setup_variables(self):
        self.mode_var = tk.StringVar(value="variable")
        self.longueurs_fixes_str = tk.StringVar(value="13000")
        self.priorite_fixe_var = tk.StringVar(value="auto")
        self.longueur_min_var = tk.IntVar(value=2500)
        self.longueur_max_var = tk.IntVar(value=13000)
        self.pas_var = tk.IntVar(value=500)
        self.marge_coupe_var = tk.IntVar(value=MARGE_COUPE_DEFAULT)
        self.valorisation_chute_var = tk.DoubleVar(value=80.0)
        self.taux_chute_mini_var = tk.DoubleVar(value=1.0)

        self.optimiser_var = tk.BooleanVar(value=True)
        self.unite_var = tk.StringVar(value="auto")
        self.unite_detectee_var = tk.StringVar(value="")

        self.methode_m3_var = tk.StringVar(value="manuel")
        self.methode_m2_var = tk.StringVar(value="manuel")  # grisé (BRUT imposé)
        self.methode_ml_var = tk.StringVar(value="manuel")

    # ----------------------
    # UI layout
    # ----------------------
    def create_interface(self):
        main = ttk.Frame(self.root); main.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)

        info = ttk.LabelFrame(main, text="Informations du projet", padding=5); info.pack(fill=tk.X, pady=(0,6))
        self.project_info_label = ttk.Label(info, text="Chargement…", foreground="blue"); self.project_info_label.pack()

        self.create_material_config_frame(main)

        cfg = ttk.LabelFrame(main, text="Configuration d'optimisation", padding=5); cfg.pack(fill=tk.BOTH, expand=True)
        self.tabs = ttk.Notebook(cfg); self.tabs.pack(fill=tk.BOTH, expand=True)
        self.create_mode_tab(self.tabs)
        self.create_params_tab(self.tabs)
        self.create_preview_tab(self.tabs)

        actions = ttk.LabelFrame(main, text="Actions", padding=6); actions.pack(fill=tk.X, pady=(6,0))
        left = ttk.Frame(actions); left.pack(fill=tk.X)
        ttk.Button(left, text="💾 Sauvegarder Config", command=self.save_config_file).pack(side=tk.LEFT, padx=3)
        ttk.Button(left, text="📂 Charger Config", command=self.load_config_file).pack(side=tk.LEFT, padx=3)
        ttk.Button(left, text="🔄 Reset", command=self.reset_to_default).pack(side=tk.LEFT, padx=3)
        self.run_btn = ttk.Button(left, text="🎯 Lancer optimisation", command=self.run_full_optimization)
        self.run_btn.pack(side=tk.LEFT, padx=10)

        self.update_project_info()

    def create_material_config_frame(self, parent):
        frame = ttk.LabelFrame(parent, text="Configuration par matériau", padding=5)
        frame.pack(fill=tk.X, pady=(0,6))

        head = ttk.Frame(frame); head.pack(fill=tk.X)
        ttk.Label(head, text="Matériau:").pack(side=tk.LEFT)
        self.material_combo = ttk.Combobox(head, state="readonly", width=45)
        self.material_combo.pack(side=tk.LEFT, padx=(8,0), fill=tk.X, expand=True)
        self.material_combo.bind("<<ComboboxSelected>>", self.on_material_selected)

        opts = ttk.LabelFrame(frame, text="Options matériau", padding=5); opts.pack(fill=tk.X, pady=(6,0))
        row1 = ttk.Frame(opts); row1.pack(fill=tk.X, pady=2)
        self.optimiser_check = ttk.Checkbutton(row1, text="Optimiser ce matériau", variable=self.optimiser_var, command=self.on_optimiser_changed)
        self.optimiser_check.pack(side=tk.LEFT)

        row2 = ttk.Frame(opts); row2.pack(fill=tk.X, pady=2)
        ttk.Label(row2, text="Unité:").pack(side=tk.LEFT)
        self.unite_info_label = ttk.Label(row2, text="", foreground="blue"); self.unite_info_label.pack(side=tk.LEFT, padx=(10,0))
        radios = ttk.Frame(row2); radios.pack(side=tk.RIGHT)
        for lbl, val in (("Auto","auto"),("m³","m3"),("m²","m2"),("ml","ml")):
            ttk.Radiobutton(radios, text=lbl, variable=self.unite_var, value=val, command=self.update_unite_info).pack(side=tk.LEFT, padx=5)

        self.create_calcul_methods_frame(opts)

        btns = ttk.Frame(frame); btns.pack(fill=tk.X, pady=(6,0))
        ttk.Button(btns, text="🔄 Actualiser", command=self.detect_materials).pack(side=tk.LEFT, padx=3)
        ttk.Button(btns, text="📊 Analyser", command=self.analyze_current_material).pack(side=tk.LEFT, padx=3)

    def create_calcul_methods_frame(self, parent):
        box = ttk.LabelFrame(parent, text="Méthode de calcul (si non optimisé)", padding=5)
        box.pack(fill=tk.X, pady=(6,0))
        grid = ttk.Frame(box); grid.pack(fill=tk.X)

        # Volume
        col_v = ttk.Frame(grid); col_v.grid(row=0, column=0, sticky="nw", padx=(0,12))
        ttk.Label(col_v, text="Volume (m³)").pack(anchor=tk.W)
        self.methode_m3_combo = ttk.Combobox(col_v, textvariable=self.methode_m3_var, state="readonly", width=20)
        self.methode_m3_combo['values'] = ["manuel", "volume_standard", "volume_liste", "volume_physique_reel"]
        self.methode_m3_combo.pack(pady=2)
        self.methode_m3_combo.bind("<<ComboboxSelected>>", self.on_method_combo_changed)

        # Surface
        col_s = ttk.Frame(grid); col_s.grid(row=0, column=1, sticky="nw", padx=(0,12))
        ttk.Label(col_s, text="Surface (m²) — BRUT imposé").pack(anchor=tk.W)
        self.methode_m2_combo = ttk.Combobox(col_s, textvariable=self.methode_m2_var, state="disabled", width=20)
        self.methode_m2_combo['values'] = ["manuel"]
        self.methode_m2_combo.set("manuel")

        # Longueur
        col_l = ttk.Frame(grid); col_l.grid(row=0, column=2, sticky="nw")
        ttk.Label(col_l, text="Longueur (m)").pack(anchor=tk.W)
        self.methode_ml_combo = ttk.Combobox(col_l, textvariable=self.methode_ml_var, state="readonly", width=20)
        self.methode_ml_combo['values'] = ["manuel", "longueur_liste", "longueur_avec_depassement"]
        self.methode_ml_combo.pack(pady=2)
        self.methode_ml_combo.bind("<<ComboboxSelected>>", self.on_method_combo_changed)

        ttk.Label(box, text="En m² non optimisé, la surface BRUTE = (2 plus grandes cotes).", foreground="gray", font=("Arial",8)).pack(anchor=tk.W, pady=(6,0))

        self.update_methods_state()

    # ----------------------
    # Callbacks
    # ----------------------
    def on_method_combo_changed(self, event=None):
        self.save_current_material_config()
        self.update_preview()

    def on_material_selected(self, event=None):
        if self.current_material is not None:
            self.save_current_material_config()  # sauvegarde l’ancien
        m = self.material_combo.get()
        if not m: return
        self.current_material = m
        self.load_material_config(m)
        self.update_unite_info()
        self.update_preview()

    def on_optimiser_changed(self):
        enable = self.optimiser_var.get()
        for i in range(self.tabs.index("end")):
            self.tabs.tab(i, state=("normal" if enable else "disabled"))
        self.update_methods_state()
        self.save_current_material_config()
        self.update_preview()

    def update_methods_state(self):
        u = (self.unite_detectee_var.get() or 'm3').lower()
        states = {'m3':'disabled', 'm2':'disabled', 'ml':'disabled'}
        if not self.optimiser_var.get():
            states[u] = 'readonly'
        self.methode_m3_combo['state'] = states['m3']
        self.methode_ml_combo['state'] = states['ml']
        self.methode_m2_combo['state'] = 'disabled'  # m² BRUT imposé

    def create_mode_tab(self, notebook):
        f = ttk.Frame(notebook); notebook.add(f, text="Mode")
        ttk.Radiobutton(f, text="Longueurs FIXES", variable=self.mode_var, value="fixe",
                        command=self.on_mode_changed).pack(anchor=tk.W, pady=2)
        ttk.Radiobutton(f, text="Longueurs VARIABLES", variable=self.mode_var, value="variable",
                        command=self.on_mode_changed).pack(anchor=tk.W, pady=2)
        self.mode_specific = ttk.Frame(f); self.mode_specific.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.update_mode_ui()

    def on_mode_changed(self):
        self.update_mode_ui()
        self.save_current_material_config()
        self.update_preview()

    def update_mode_ui(self):
        for w in self.mode_specific.winfo_children():
            w.destroy()
        if self.mode_var.get() == 'fixe':
            fr = ttk.LabelFrame(self.mode_specific, text="Longueurs fixes", padding=5); fr.pack(fill=tk.BOTH, expand=True)
            r = ttk.Frame(fr); r.pack(fill=tk.X, pady=2)
            ttk.Label(r, text="Longueurs (mm):").pack(side=tk.LEFT)
            ttk.Entry(r, textvariable=self.longueurs_fixes_str, width=32).pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(10,0))
            ttk.Label(fr, text="ex: 4000, 6000, 13000", foreground="gray", font=("Arial",8)).pack(anchor=tk.W)
            pr = ttk.LabelFrame(fr, text="Priorité", padding=4); pr.pack(fill=tk.X, pady=6)
            for v,t in (("auto","Auto"),("petit","Petites d'abord"),("grand","Grandes d'abord")):
                ttk.Radiobutton(pr, text=t, value=v, variable=self.priorite_fixe_var).pack(anchor=tk.W)
        else:
            fr = ttk.LabelFrame(self.mode_specific, text="Longueurs variables", padding=5); fr.pack(fill=tk.BOTH, expand=True)
            p = ttk.Frame(fr); p.pack(fill=tk.X)
            ttk.Label(p, text="Longueur mini (mm):").grid(row=0,column=0,sticky=tk.W); ttk.Spinbox(p, from_=1000,to=10000, textvariable=self.longueur_min_var, width=10).grid(row=0,column=1,padx=8)
            ttk.Label(p, text="Longueur maxi (mm):").grid(row=1,column=0,sticky=tk.W); ttk.Spinbox(p, from_=5000,to=20000, textvariable=self.longueur_max_var, width=10).grid(row=1,column=1,padx=8)
            ttk.Label(p, text="Pas (mm):").grid(row=2,column=0,sticky=tk.W); ttk.Spinbox(p, from_=100,to=1000, textvariable=self.pas_var, width=10).grid(row=2,column=1,padx=8)

    def create_params_tab(self, notebook):
        f = ttk.Frame(notebook); notebook.add(f, text="Paramètres")
        cut = ttk.LabelFrame(f, text="Paramètres de coupe", padding=5); cut.pack(fill=tk.X, padx=5, pady=5)
        r = ttk.Frame(cut); r.pack(fill=tk.X)
        ttk.Label(r, text="Marge de coupe (mm):").pack(side=tk.LEFT)
        ttk.Spinbox(r, from_=50,to=200, textvariable=self.marge_coupe_var, width=10).pack(side=tk.RIGHT)
        eco = ttk.LabelFrame(f, text="Paramètres économiques", padding=5); eco.pack(fill=tk.X, padx=5, pady=5)
        r1 = ttk.Frame(eco); r1.pack(fill=tk.X, pady=2)
        ttk.Label(r1, text="Valorisation chute (€/unité):").pack(side=tk.LEFT)
        ttk.Spinbox(r1, from_=0.0,to=200.0, increment=5.0, textvariable=self.valorisation_chute_var, width=10).pack(side=tk.RIGHT)
        r2 = ttk.Frame(eco); r2.pack(fill=tk.X, pady=2)
        ttk.Label(r2, text="Taux chute mini (%):").pack(side=tk.LEFT)
        ttk.Spinbox(r2, from_=0.0,to=20.0, increment=0.5, textvariable=self.taux_chute_mini_var, width=10).pack(side=tk.RIGHT)

    def create_preview_tab(self, notebook):
        f = ttk.Frame(notebook); notebook.add(f, text="Aperçu")
        self.preview_text = tk.Text(f, height=12, width=80, wrap=tk.WORD)
        sb = ttk.Scrollbar(f, orient=tk.VERTICAL, command=self.preview_text.yview)
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); sb.pack(side=tk.RIGHT, fill=tk.Y)
        self.preview_text.config(yscrollcommand=sb.set)
        ttk.Button(f, text="🔄 Mettre à jour", command=self.update_preview).pack(pady=6)

    # ----------------------
    # Data
    # ----------------------
    def detect_materials(self):
        try:
            element_ids = ec.get_active_identifiable_element_ids()
            mats, elems = set(), []
            for eid in element_ids:
                try:
                    mat = ac.get_element_material_name(eid).strip()
                    if not mat: continue
                    mats.add(mat)
                    elems.append({'eid':eid, 'materiau':mat,
                                  'longueur': safe_float(gc.get_length(eid)),
                                  'largeur': round(_get_list_width(eid)),
                                  'hauteur': round(_get_list_height(eid))})
                except Exception as e:
                    log_message(f"Element {eid}: {e}", "WARNING")
            self.materiaux_detectes = sorted(list(mats))
            self.elements_data = elems
            self.material_combo['values'] = self.materiaux_detectes
            if self.materiaux_detectes and not self.current_material:
                self.material_combo.set(self.materiaux_detectes[0]); self.on_material_selected()
            self.update_project_info()
        except Exception as e:
            log_message(f"Détection: {e}", "ERROR"); messagebox.showerror("Erreur", f"Détection matériaux:\n{e}")

    def update_project_info(self):
        self.project_info_label.config(text=f"{len(self.elements_data)} éléments • {len(self.materiaux_detectes)} matériaux détectés")
        self.run_btn.config(state=("normal" if self.elements_data else "disabled"))

    def update_unite_info(self):
        if not self.current_material: return
        choice = self.unite_var.get()
        if choice == 'auto':
            u = get_material_unit(self.current_material)
            self.unite_detectee_var.set(u); self.unite_info_label.config(text=f"Détectée: {u.upper()}")
        else:
            self.unite_detectee_var.set(choice); self.unite_info_label.config(text=f"Forcée: {choice.upper()}")
        self.save_current_material_config(); self.update_preview(); self.update_methods_state()

    def get_default_config(self, material):
        unite_auto = get_material_unit(material)
        up = (material or '').upper()

        # défaut m³ : manuel ; pour cintrés -> volume_physique_reel
        methode_m3_defaut = "manuel"
        if "CINTRE" in up or "CINTR" in up:
            methode_m3_defaut = "volume_physique_reel"

        methode_m2_defaut = "manuel"  # m² BRUT imposé

        base = {
            'optimiser': True, 'unite': 'auto', 'unite_detectee': unite_auto,
            'mode': 'variable', 'longueurs_fixes': [13000], 'priorite_fixe': 'auto',
            'longueur_min': 2500, 'longueur_max': 13000, 'pas': 500,
            'marge_coupe': MARGE_COUPE_DEFAULT, 'valorisation_chute': 80.0, 'taux_chute_mini': 1.0,
            'methode_m3': methode_m3_defaut, 'methode_m2': methode_m2_defaut, 'methode_ml': "manuel"
        }
        if is_materiau_13m(material): base['mode'] = 'fixe'
        return base

    def save_current_material_config(self):
        if not self.current_material: return
        try:
            liste_fixes = []
            if self.longueurs_fixes_str.get().strip():
                for chunk in self.longueurs_fixes_str.get().split(','):
                    chunk = chunk.strip()
                    if chunk: liste_fixes.append(int(chunk))
            self.materiaux_configs[self.current_material] = {
                'optimiser': self.optimiser_var.get(),
                'unite': self.unite_var.get(),
                'unite_detectee': self.unite_detectee_var.get(),
                'mode': self.mode_var.get(),
                'longueurs_fixes': liste_fixes,
                'priorite_fixe': self.priorite_fixe_var.get(),
                'longueur_min': self.longueur_min_var.get(),
                'longueur_max': self.longueur_max_var.get(),
                'pas': self.pas_var.get(),
                'marge_coupe': self.marge_coupe_var.get(),
                'valorisation_chute': self.valorisation_chute_var.get(),
                'taux_chute_mini': self.taux_chute_mini_var.get(),
                'methode_m3': self.methode_m3_var.get(),
                'methode_m2': self.methode_m2_var.get(),
                'methode_ml': self.methode_ml_var.get(),
            }
        except ValueError as e:
            messagebox.showerror("Erreur", f"Valeurs invalides: {e}")

    def load_material_config(self, material):
        cfg = self.materiaux_configs.get(material) or self.get_default_config(material)
        self.materiaux_configs[material] = cfg
        self.optimiser_var.set(cfg.get('optimiser', True))
        self.unite_var.set(cfg.get('unite', 'auto')); self.unite_detectee_var.set(cfg.get('unite_detectee', get_material_unit(material)))
        self.mode_var.set(cfg.get('mode','variable'))
        self.longueurs_fixes_str.set(", ".join(map(str, cfg.get('longueurs_fixes', [13000]))))
        self.priorite_fixe_var.set(cfg.get('priorite_fixe','auto'))
        self.longueur_min_var.set(cfg.get('longueur_min',2500)); self.longueur_max_var.set(cfg.get('longueur_max',13000)); self.pas_var.set(cfg.get('pas',500))
        self.marge_coupe_var.set(cfg.get('marge_coupe',MARGE_COUPE_DEFAULT))
        self.valorisation_chute_var.set(cfg.get('valorisation_chute',80.0))
        self.taux_chute_mini_var.set(cfg.get('taux_chute_mini',1.0))
        self.methode_m3_var.set(cfg.get('methode_m3','manuel'))
        self.methode_m2_var.set(cfg.get('methode_m2','manuel'))
        self.methode_ml_var.set(cfg.get('methode_ml','manuel'))
        self.update_mode_ui(); self.on_optimiser_changed()

    def load_default_configs(self):
        for m in self.materiaux_detectes:
            if m not in self.materiaux_configs:
                self.materiaux_configs[m] = self.get_default_config(m)

    def analyze_current_material(self):
        if not self.current_material:
            messagebox.showwarning("Attention", "Aucun matériau sélectionné"); return
        material_elements = [e for e in self.elements_data if e['materiau'] == self.current_material]
        if not material_elements:
            messagebox.showinfo("Info", f"Aucun élément pour {self.current_material}"); return
        longueurs = [e['longueur'] for e in material_elements]
        unite_detectee = get_material_unit(self.current_material)
        if unite_detectee == 'ml':
            qtot = sum(e['longueur'] for e in material_elements) / 1000
            qstr = f"Quantité totale: {qtot:.2f} m"
        elif unite_detectee == 'm2':
            stot = sum((e['largeur']*e['hauteur'])/1_000_000 for e in material_elements)
            qstr = f"Surface totale (approx.): {stot:.2f} m²"
        else:
            vtot = sum((e['largeur']*e['hauteur']*e['longueur'])/1_000_000_000 for e in material_elements)
            qstr = f"Volume total: {vtot:.4f} m³"
        stats = f"""📊 ANALYSE - {self.current_material}
Nombre d'éléments: {len(material_elements)}
Unité détectée: {unite_detectee.upper()}
Longueurs: {min(longueurs):.0f} - {max(longueurs):.0f} mm
{qstr}"""
        messagebox.showinfo("Analyse matériau", stats)

    def update_preview(self):
        if not self.current_material:
            self.preview_text.delete(1.0, tk.END); self.preview_text.insert(tk.END, "Sélectionnez un matériau."); return
        self.save_current_material_config()
        cfg = self.materiaux_configs.get(self.current_material, {})
        ueff = cfg.get('unite_detectee','m3')
        optimiser = cfg.get('optimiser', True)
        preview = [f"🎯 {self.current_material}", ""]
        preview.append(f"Optimisé: {'Oui' if optimiser else 'Non'}")
        preview.append(f"Unité: {ueff.upper()}")
        if not optimiser and ueff=='m2':
            preview.append("Méthode surface: BRUT (2 plus grandes cotes)")
        if not optimiser and ueff=='m3':
            preview.append(f"Méthode volume: {cfg.get('methode_m3','manuel')}")
        if not optimiser and ueff=='ml':
            preview.append(f"Méthode longueur: {cfg.get('methode_ml','manuel')}")
        self.preview_text.delete(1.0, tk.END); self.preview_text.insert(tk.END, "\n".join(preview))

    # --- Fichiers config ---
    def save_config_file(self):
        self.save_current_material_config()
        filename = filedialog.asksaveasfilename(title="Sauvegarder configuration", defaultextension=".json",
                                                filetypes=[("Configuration JSON","*.json")])
        if not filename: return
        try:
            data = {'version':'5.2_brut_m2','date':datetime.now().isoformat(),'materiaux':self.materiaux_configs}
            with open(filename,'w',encoding='utf-8') as f: json.dump(data, f, indent=2, ensure_ascii=False)
            messagebox.showinfo("Succès", f"Config sauvegardée:\n{filename}")
        except Exception as e:
            messagebox.showerror("Erreur", f"Sauvegarde: {e}")

    def load_config_file(self):
        filename = filedialog.askopenfilename(title="Charger configuration", filetypes=[("Configuration JSON","*.json")])
        if not filename: return
        try:
            with open(filename,'r',encoding='utf-8') as f: data = json.load(f)
            if 'materiaux' in data: self.materiaux_configs.update(data['materiaux']); self.load_material_config(self.current_material)
            messagebox.showinfo("Succès", f"Configuration chargée.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Chargement: {e}")

    def reset_to_default(self):
        if not self.current_material: return
        if messagebox.askyesno("Confirmation", f"Réinitialiser {self.current_material} ?"):
            self.materiaux_configs[self.current_material] = self.get_default_config(self.current_material)
            self.load_material_config(self.current_material)

    # --- Optimisation ---
    def preview_optimization(self):
        try:
            self.save_current_material_config()
            res = ["Aperçu optimisation", "", f"Matériaux configurés: {len(self.materiaux_configs)}", f"Éléments: {len(self.elements_data)}"]
            messagebox.showinfo("Aperçu", "\n".join(res))
        except Exception as e:
            messagebox.showerror("Erreur", f"Aperçu: {e}")

    def run_full_optimization(self):
        try:
            self.save_current_material_config()
            if not self.elements_data:
                messagebox.showwarning("Attention", "Aucun élément"); return
            self.execute_optimization()
        except Exception as e:
            messagebox.showerror("Erreur", f"Lancement: {e}")

    def execute_optimization(self):
        log_message("Début optimisation...")
        info_mats, table_cmd, table_barres, = optimiser_avec_unites(self.materiaux_configs, self.elements_data)
        self.generer_excel(info_mats, table_cmd, table_barres)
        messagebox.showinfo("Terminé", "Optimisation terminée et Excel généré.")
        self.root.destroy()

    # --- Excel ---
    def generer_excel(self, info_materiaux, tableau_commandes, tableau_barres_detaille):
        wb = Workbook()
        ws_recap = wb.active; ws_recap.title = "Récapitulatif"
        self.remplir_feuille_et_formater(
            ws_recap,
            ["Matériau","Unité","Qte opti","Qte brut","Prix U (€)","Prix Tot (€)","Taux chute (%)","Optimisé"],
            [[
                mat,
                data.get('unite','m3').upper().replace('M2','M²').replace('M3','M³'),
                round(data.get('volume_barre',0.0),4),
                round(data.get('volume_utilise',0.0),4),
                round(data.get('prix_unitaire',0.0),2),
                round(data.get('prix_unitaire',0.0)*data.get('volume_barre',0.0),2),
                (round(100*(1 - (data.get('volume_utilise',0.0)/(data.get('volume_barre',1.0) or 1.0))),2) if data.get('optimise',True) else 0.0),
                "Oui" if data.get('optimise',True) else "Non"
            ] for mat, data in info_materiaux.items()],
            "TableRecap"
        )

        ws_cmd = wb.create_sheet("Commandes")
        self.remplir_feuille_et_formater(
            ws_cmd,
            ["Matériau","Largeur (mm)","Hauteur (mm)","Longueurs","Total barres","Taux chute (%)","Quantité","Prix U (€)","Prix total (€)"],
            tableau_commandes,
            "TableCommandes"
        )

        ws_barres = wb.create_sheet("Barres détaillées")
        self.remplir_feuille_et_formater(
            ws_barres,
            ["ID Barre","Matériau","Largeur (mm)","Hauteur (mm)","Longueur barre (mm)","Nb pièces","Pièces","Chute (mm)","Taux chute (%)","Quantité","Prix (€)"],
            tableau_barres_detaille,
            "TableBarres"
        )

        num_devis = uc.get_project_user_attribute(1).strip().replace(" ","_")
        client = uc.get_project_number().strip().replace(" ","_")
        nom_fichier = f"{num_devis}-{client}-optimisation_v5-2_brut_m2.xlsx"
        output_path = os.path.join(os.path.expanduser("~"), "Desktop", nom_fichier)
        wb.save(output_path)
        try:
            subprocess.Popen(['start', '', output_path], shell=True)
        except Exception:
            pass
        log_message(f"Excel créé: {output_path}")

    def remplir_feuille_et_formater(self, ws, titres, donnees, table_name):
        ws.append(titres)
        for row in donnees:
            ws.append(row)
        if donnees:
            total_row = ["TOTAL"] + ["" for _ in range(len(titres)-1)]
            last_data_row = len(donnees) + 1
            for i in range(1, len(titres)):
                col = get_column_letter(i+1)
                total_row[i] = f"=SUBTOTAL(109,{col}2:{col}{last_data_row})"
            ws.append(total_row)
        last_row = ws.max_row; last_col = ws.max_column
        if last_row > 1:
            tab = Table(displayName=table_name, ref=f"A1:{get_column_letter(last_col)}{last_row}")
            tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(tab)
        for c in range(1, last_col+1):
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[get_column_letter(c)])
            ws.column_dimensions[get_column_letter(c)].width = max_len + 2

# =========================================================
# Optimisation
# =========================================================

def optimiser_avec_unites(materiaux_configs, elements_data):
    info_materiaux: Dict[str,dict] = {}
    tableau_commandes: List[list] = []
    tableau_barres_detaille: List[list] = []
    barre_global_id = 1

    groupes = defaultdict(list)
    for e in elements_data:
        groupes[e['materiau']].append(e)

    for mat_name, elements in groupes.items():
        cfg = materiaux_configs.get(mat_name, {})
        if not cfg.get('optimiser', True):
            traiter_materiau_non_optimise(mat_name, elements, info_materiaux, tableau_commandes, tableau_barres_detaille, cfg)
            continue

        unite = cfg.get('unite_detectee', 'm3')
        longueurs_barres = generer_longueurs_materiau(mat_name, cfg)
        marge = cfg.get('marge_coupe', MARGE_COUPE_DEFAULT)
        barre_global_id = optimiser_materiau_avec_unite(elements, mat_name, unite, longueurs_barres, marge,
                                                        info_materiaux, tableau_commandes, tableau_barres_detaille, barre_global_id)
    return info_materiaux, tableau_commandes, tableau_barres_detaille

def traiter_materiau_non_optimise(mat_name, elements, info_materiaux, tableau_commandes, tableau_barres_detaille, config=None):
    # init fiche matériau
    if mat_name not in info_materiaux:
        try:
            mat_id = mc.get_material_id(mat_name); prix_u = safe_float(mc.get_price(mat_id))
        except Exception:
            prix_u = 0.0
        info_materiaux[mat_name] = {
            'volume_utilise': 0.0, 'volume_barre': 0.0, 'prix_unitaire': prix_u,
            'unite': (config.get('unite_detectee') if config else get_material_unit(mat_name)),
            'optimise': False
        }

    unite_eff = info_materiaux[mat_name]['unite']  # m3 | m2 | ml
    # méthode selon unité effective
    if unite_eff == 'm3':
        method = (config.get('methode_m3') if config else None) or 'volume_physique_reel'
    elif unite_eff == 'ml':
        method = (config.get('methode_ml') if config else None) or 'manuel'
    else:
        method = 'manuel'  # m² = BRUT imposé

    log_message(f"Non optimisé {mat_name} (unite={unite_eff}) method={method}", "DEBUG")

    for element in elements:
        eid = element['eid']
        q = calculate_quantity_with_cadwork_method(eid, unite_eff, method)
        info_materiaux[mat_name]['volume_utilise'] += q
        info_materiaux[mat_name]['volume_barre'] += q
        ac.set_user_attribute([eid], 12, "NON_OPTIMISE")
        ac.set_user_attribute([eid], 13, "0.0 %")

    qtot = info_materiaux[mat_name]['volume_barre']; prix_u = info_materiaux[mat_name]['prix_unitaire']
    tableau_commandes.append([mat_name,"N/A","N/A","NON OPTIMISE",len(elements),0.0,round(qtot,4),prix_u,round(qtot*prix_u,2)])

def optimiser_materiau_avec_unite(elements, mat_name, unite, longueurs_barres, marge_coupe,
                                  info_materiaux, tableau_commandes, tableau_barres_detaille, barre_global_id):
    if mat_name not in info_materiaux:
        try:
            mat_id = mc.get_material_id(mat_name); prix_u = safe_float(mc.get_price(mat_id))
        except Exception:
            prix_u = 0.0
        info_materiaux[mat_name] = {'volume_utilise':0.0,'volume_barre':0.0,'prix_unitaire':prix_u,'unite':unite,'optimise':True}

    groupes_sections = defaultdict(list)
    for e in elements:
        groupes_sections[(e['largeur'], e['hauteur'])].append(e['eid'])

    regroup_cmd = defaultdict(lambda: {'longueurs': defaultdict(int),'taux_chute':[],'quantite_total':0,'prix_total':0})

    for (largeur, hauteur), eids in groupes_sections.items():
        eids.sort(key=lambda eid: gc.get_length(eid), reverse=True)
        barres: List[List[int]] = []
        for eid in eids:
            piece_L = safe_float(gc.get_length(eid))
            best, min_perte = None, max(longueurs_barres)+1
            for barre in barres:
                nb = len(barre)
                L_sans_marge = sum(gc.get_length(x) for x in barre)
                marge_tot = nb * marge_coupe
                occ = L_sans_marge + marge_tot
                cand = [l for l in longueurs_barres if l >= occ + piece_L]
                if cand:
                    Lp = min(cand)
                    perte = Lp - (occ + piece_L)
                    if perte < min_perte:
                        min_perte, best = perte, barre
            if best is not None: best.append(eid)
            else: barres.append([eid])

        for barre in barres:
            nb = len(barre)
            L_sans_marge = sum(gc.get_length(x) for x in barre)
            marge_tot = (nb-1)*marge_coupe if nb>1 else 0
            occ = L_sans_marge + marge_tot
            possibles = [l for l in longueurs_barres if l >= occ]
            L_finale = min(possibles) if possibles else max(longueurs_barres)
            chute = max(L_finale - occ, 0)

            quantite_barre = calculate_quantity_by_unit(largeur, hauteur, L_finale, unite)
            info_materiaux[mat_name]['volume_barre'] += quantite_barre
            prix_u = info_materiaux[mat_name]['prix_unitaire']; prix_barre = prix_u * quantite_barre

            q_pieces = [calculate_quantity_by_unit(largeur, hauteur, gc.get_length(x), unite) for x in barre]
            somme_q = sum(q_pieces)
            taux = (quantite_barre - somme_q)/quantite_barre if quantite_barre>0 else 0

            tableau_barres_detaille.append([barre_global_id, mat_name, largeur, hauteur, int(L_finale), nb,
                                            " | ".join(map(str,barre)), int(chute), round(taux*100,2),
                                            round(quantite_barre,4), round(prix_barre,2)])

            regroup_cmd[(mat_name, largeur, hauteur)]['longueurs'][int(L_finale)] += 1
            regroup_cmd[(mat_name, largeur, hauteur)]['taux_chute'].append(taux*100)
            regroup_cmd[(mat_name, largeur, hauteur)]['quantite_total'] += quantite_barre
            regroup_cmd[(mat_name, largeur, hauteur)]['prix_total'] += prix_barre

            for eid, qp in zip(barre, q_pieces):
                ac.set_user_attribute([eid], 12, str(barre_global_id))
                ac.set_user_attribute([eid], 13, f"{round(taux*100,2)} %")
                info_materiaux[mat_name]['volume_utilise'] += qp

            barre_global_id += 1

    for (mat_key, largeur, hauteur), data in regroup_cmd.items():
        taux_moy = (sum(data['taux_chute'])/len(data['taux_chute'])) if data['taux_chute'] else 0
        prix_u = info_materiaux[mat_name]['prix_unitaire']
        longueurs_str = " | ".join([f"{L}mm x{qty}" for L,qty in sorted(data['longueurs'].items())])
        tableau_commandes.append([mat_key, largeur, hauteur, longueurs_str, sum(data['longueurs'].values()),
                                  round(taux_moy,2), round(data['quantite_total'],4), round(prix_u,2),
                                  round(data['prix_total'],2)])
    return barre_global_id

def generer_longueurs_materiau(material_name, config):
    mode = config.get('mode','variable')
    if mode == 'fixe':
        lengths = sorted(config.get('longueurs_fixes',[13000]))
        p = config.get('priorite_fixe','auto')
        return sorted(lengths) if p!='grand' else sorted(lengths, reverse=True)
    min_l = config.get('longueur_min',2500); max_l = config.get('longueur_max',13000); pas = config.get('pas',500)
    lengths = [l for l in range(min_l, max_l+1, pas)]
    if 7000 not in lengths and min_l <= 7000 <= max_l: lengths.insert(0,7000)
    return lengths

# =========================================================
# Main
# =========================================================

def main():
    log_message("Optimisation Scierie v5.2 — démarrage")
    try:
        _ = ec.get_active_identifiable_element_ids()
    except Exception as e:
        messagebox.showerror("Erreur Cadwork", f"Connexion Cadwork impossible:\\n{e}")
        return
    app = OptimizationConfigUI()
    app.root.mainloop()

if __name__ == '__main__':
    main()
