import sys
import os
sys.path.append(r"C:\cadwork\libs")

import cadwork
import geometry_controller as gc
import attribute_controller as ac
import element_controller as ec
import material_controller as mc
import utility_controller as uc
from datetime import datetime
from collections import defaultdict
import subprocess
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading

class DevisGroupeInterface:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Génération Devis par Groupe - Cadwork")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')
        
        # Constantes
        self.VALORISATION_CHUTE = 80
        self.prix_traitement = {"CL2": 25.5, "CL3": 215, "CL2 INCOLORE": 50}
        self.prix_faconnage = {"T1_V": 224.4, "T16_L": 2.34, "T2_V": 243.1, "T3_V": 320, "T4_V": 263.5}
        self.prix_prestation = {
            "MO1_S": 17, "MO2_S": 28.9, "MO3_S": 22.95, "MO4_S": 19.55,
            "MO5_V": 219.3, "MO6_V": 85, "MO7_V": 119, "MO8_V": 35
        }
        
        # Variables
        self.element_ids = []
        self.should_stop = False
        self.elements_traites = 0
        self.groupes_detectes = 0
        self.ventilation_data = {}
        self.details_par_groupe = {}
        
        self.setup_ui()
        
    def setup_ui(self):
        # Style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Titre principal
        title_frame = tk.Frame(self.root, bg='#27ae60', height=80)
        title_frame.pack(fill='x', padx=0, pady=0)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="📊 GÉNÉRATION DEVIS PAR GROUPE", 
                              font=('Arial', 18, 'bold'), fg='white', bg='#27ae60')
        title_label.pack(expand=True)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Zone d'information
        info_frame = tk.LabelFrame(main_frame, text="ℹ️ Configuration", font=('Arial', 12, 'bold'), 
                                 bg='#f0f0f0', fg='#27ae60')
        info_frame.pack(fill='x', pady=(0, 10))
        
        info_text = tk.Label(info_frame, 
                           text="Génère un devis détaillé par groupe et sous-groupe :\n"
                                "• Groupe principal : Attribut utilisateur 3 (CDPGF)\n"
                                "• Sous-groupe : Attribut utilisateur 16\n"
                                "• Deux feuilles Excel : Analyse détaillée + Présentation hiérarchique",
                           font=('Arial', 10), bg='#f0f0f0', fg='#2c3e50', justify='left')
        info_text.pack(pady=10, padx=10)
        
        # Zone de statut
        status_frame = tk.LabelFrame(main_frame, text="📊 Statut", font=('Arial', 12, 'bold'), 
                                   bg='#f0f0f0', fg='#27ae60')
        status_frame.pack(fill='x', pady=(0, 10))
        
        self.status_label = tk.Label(status_frame, text="⏳ En attente...", 
                                   font=('Arial', 11), bg='#f0f0f0', fg='#2c3e50')
        self.status_label.pack(pady=10)
        
        # Statistiques
        stats_frame = tk.Frame(status_frame, bg='#f0f0f0')
        stats_frame.pack(fill='x', padx=20, pady=(0, 10))
        
        self.stats_label = tk.Label(stats_frame, text="📈 Statistiques : 0 éléments traités | 0 groupes détectés", 
                                  font=('Arial', 10), bg='#f0f0f0', fg='#2c3e50')
        self.stats_label.pack()
        
        # Barre de progression
        self.progress = ttk.Progressbar(status_frame, mode='determinate')
        self.progress.pack(fill='x', padx=20, pady=(0, 10))
        
        # Zone de résultats
        results_frame = tk.LabelFrame(main_frame, text="📋 Journal de génération", font=('Arial', 12, 'bold'), 
                                    bg='#f0f0f0', fg='#27ae60')
        results_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # Zone de texte avec scroll
        self.text_area = scrolledtext.ScrolledText(results_frame, height=15, width=80, 
                                                 font=('Courier', 9), bg='#ffffff', 
                                                 fg='#2c3e50', wrap=tk.WORD)
        self.text_area.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Zone des boutons
        button_frame = tk.Frame(main_frame, bg='#f0f0f0')
        button_frame.pack(fill='x', pady=(10, 0))
        
        self.start_button = tk.Button(button_frame, text="🚀 GÉNÉRER LE DEVIS", 
                                    command=self.start_generation, font=('Arial', 12, 'bold'),
                                    bg='#27ae60', fg='white', height=2, cursor='hand2')
        self.start_button.pack(side='left', padx=(0, 10))
        
        self.stop_button = tk.Button(button_frame, text="⏹️ ARRÊTER", 
                                   command=self.stop_generation, font=('Arial', 12, 'bold'),
                                   bg='#e74c3c', fg='white', height=2, cursor='hand2',
                                   state='disabled')
        self.stop_button.pack(side='left', padx=(0, 10))
        
        self.close_button = tk.Button(button_frame, text="🚪 FERMER", 
                                    command=self.close_application, font=('Arial', 12, 'bold'),
                                    bg='#95a5a6', fg='white', height=2, cursor='hand2')
        self.close_button.pack(side='right')
        
        # Gestionnaire de fermeture de fenêtre
        self.root.protocol("WM_DELETE_WINDOW", self.close_application)
        
    def log(self, message, color='#2c3e50'):
        """Ajoute un message dans la zone de texte"""
        self.text_area.config(state='normal')
        self.text_area.insert(tk.END, f"{message}\n")
        self.text_area.config(state='disabled')
        self.text_area.see(tk.END)
        self.root.update()
        
    def update_status(self, message):
        """Met à jour le statut"""
        self.status_label.config(text=message)
        self.root.update()
        
    def update_stats(self):
        """Met à jour les statistiques"""
        stats_text = f"📈 Statistiques : {self.elements_traites} éléments traités | {self.groupes_detectes} groupes détectés"
        self.stats_label.config(text=stats_text)
        self.root.update()
        
    def close_application(self):
        """Ferme l'application proprement"""
        if hasattr(self, 'should_stop'):
            self.should_stop = True
        self.root.destroy()
        
    def start_generation(self):
        """Lance la génération dans un thread séparé"""
        self.should_stop = False
        self.elements_traites = 0
        self.groupes_detectes = 0
        
        self.start_button.config(state='disabled')
        self.stop_button.config(state='normal')
        
        self.text_area.config(state='normal')
        self.text_area.delete(1.0, tk.END)
        self.text_area.config(state='disabled')
        
        # Lancer dans un thread pour ne pas bloquer l'interface
        thread = threading.Thread(target=self.run_generation)
        thread.daemon = True
        thread.start()
        
    def stop_generation(self):
        """Arrête la génération"""
        self.should_stop = True
        self.update_status("🛑 Arrêt en cours...")
        
    def safe_float(self, val, default=0.0):
        try:
            return float(str(val).replace('%', '').replace(',', '.'))
        except (ValueError, TypeError):
            return default

    def get_prix_matiere(self, eid, mat_name):
        prix_att = self.safe_float(ac.get_user_attribute(eid, 6))
        return prix_att if prix_att > 0 else self.safe_float(mc.get_price(mc.get_material_id(mat_name)))

    def run_generation(self):
        """Fonction principale de génération"""
        try:
            self.update_status("🔍 Récupération des éléments...")
            self.element_ids = ec.get_active_identifiable_element_ids()
            
            if not self.element_ids:
                self.log("❌ Aucun élément sélectionné !")
                messagebox.showwarning("Aucun élément", "Veuillez sélectionner des éléments avant de générer le devis.")
                return
                
            self.log(f"🔍 Traitement de {len(self.element_ids)} élément(s) pour génération du devis")
            self.log("=" * 80)
            
            # Configuration de la barre de progression
            self.progress.config(maximum=len(self.element_ids), value=0)
            
            # Analyse des éléments
            self.update_status("⚙️ Analyse des éléments...")
            ventilation_groupes = self.analyze_elements()
            
            if self.should_stop:
                return
                
            # Génération du fichier Excel
            self.update_status("📊 Génération du fichier Excel...")
            self.generate_excel(ventilation_groupes)
            
            if not self.should_stop:
                self.log("=" * 80)
                self.log("🎯 GÉNÉRATION TERMINÉE !")
                self.log(f"📊 Résumé : {self.elements_traites} éléments traités")
                self.log(f"📁 {self.groupes_detectes} groupes détectés")
                
                self.update_status("✅ Génération terminée avec succès !")
                
                # Popup de fin
                messagebox.showinfo("Génération terminée", 
                                  f"Devis généré avec succès !\n\n"
                                  f"📊 {self.elements_traites} éléments traités\n"
                                  f"📁 {self.groupes_detectes} groupes détectés")
            else:
                self.update_status("🛑 Génération arrêtée")
                
        except Exception as e:
            self.log(f"❌ ERREUR GÉNÉRALE : {str(e)}")
            self.update_status(f"❌ Erreur : {str(e)}")
            messagebox.showerror("Erreur", f"Une erreur est survenue :\n{str(e)}")
            
        finally:
            self.start_button.config(state='normal')
            self.stop_button.config(state='disabled')

    def analyze_elements(self):
        """Analyse les éléments et lit les prix déjà calculés par groupe/sous-groupe"""
        ventilation_groupes = defaultdict(lambda: defaultdict(float))
        details_par_groupe = defaultdict(lambda: {
            'materiaux': defaultdict(lambda: {'qte': 0, 'prix': 0}),
            'traitements': defaultdict(lambda: {'qte': 0, 'prix': 0}),
            'faconnages': defaultdict(lambda: {'qte': 0, 'prix': 0}),
            'prestations': defaultdict(lambda: {'qte': 0, 'prix': 0})
        })
        
        for i, eid in enumerate(self.element_ids):
            if self.should_stop:
                break
                
            try:
                # Mise à jour de la progression
                if i % 10 == 0:
                    self.update_status(f"⚙️ Analyse élément {i+1}/{len(self.element_ids)}...")
                
                # NOUVEAU : Utiliser attributs 3 et 16 pour le groupement
                groupe_principal = ac.get_user_attribute(eid, 3) or "Non défini"
                sous_groupe = ac.get_user_attribute(eid, 16) or "Non défini"
                groupe_complet = f"{groupe_principal} | {sous_groupe}"
                
                # OPTIMISATION : Lire directement les prix déjà calculés dans les attributs
                prix_matiere_piece = self.safe_float(ac.get_user_attribute(eid, 8))  # Attribut 8 : Prix matière pièce
                prix_achat_piece = self.safe_float(ac.get_user_attribute(eid, 15))   # Attribut 15 : Prix achat
                prix_faconnage_piece = self.safe_float(ac.get_user_attribute(eid, 9))   # Attribut 9 : Prix façonnage
                prix_traitement_piece = self.safe_float(ac.get_user_attribute(eid, 7))  # Attribut 7 : Prix traitement
                prix_prestation_piece = self.safe_float(ac.get_user_attribute(eid, 10)) # Attribut 10 : Prix prestation
                
                # Récupération des données spécifiques pour le détail
                mat_name = ac.get_element_material_name(eid).strip()
                traitement_code = (ac.get_user_attribute(eid, 1) or "").strip()
                sku_faconnage = (ac.get_sku(eid) or "").strip()
                presta_code = (ac.get_user_attribute(eid, 2) or "").strip()
                
                # Calcul des quantités pour les statistiques
                taux_chute = self.safe_float(ac.get_user_attribute(eid, 13)) / 100.0 if ac.get_user_attribute(eid, 13) else 0.0
                longueur_utile = gc.get_length(eid)
                longueur_utile_m = longueur_utile / 1000
                largeur_brute = round(gc.get_list_width(eid))
                hauteur_brute = round(gc.get_list_height(eid))
                
                # Quantités selon type de matériau
                if mat_name.endswith("_L"):
                    qte_utile = longueur_utile_m
                    qte_achetee = qte_utile / (1.0 - taux_chute) if taux_chute < 1.0 else qte_utile
                    qte_chute_piece = qte_achetee - qte_utile
                    montant_chute_val = 0.0  # Pas de valorisation des chutes pour les matériaux linéaires
                    unite_matiere = "ml"
                else:
                    vol_utile = largeur_brute * hauteur_brute * longueur_utile / 1_000_000_000
                    qte_utile = vol_utile
                    qte_achetee = qte_utile / (1.0 - taux_chute) if taux_chute < 1.0 else qte_utile
                    qte_chute_piece = qte_achetee - qte_utile
                    montant_chute_val = qte_chute_piece * self.VALORISATION_CHUTE
                    unite_matiere = "m³"

                # Surfaces et volumes pour prestations (pour statistiques)
                surface_presta = 0.0
                volume_presta = 0.0
                if presta_code:
                    if presta_code.endswith("_S"):
                        surface_presta = gc.get_element_reference_face_area(eid) / 1_000_000
                        qte_presta = surface_presta
                        unite_presta = "m²"
                    else:
                        volume_presta = largeur_brute * hauteur_brute * longueur_utile / 1_000_000_000
                        qte_presta = volume_presta
                        unite_presta = "m³"
                else:
                    qte_presta = 0
                    unite_presta = ""

                # === ACCUMULATION DES DÉTAILS SPÉCIFIQUES ===
                details = details_par_groupe[groupe_complet]
                
                # Matériaux
                if prix_achat_piece > 0:
                    details['materiaux'][mat_name]['qte'] += qte_achetee
                    details['materiaux'][mat_name]['prix'] += prix_achat_piece
                    details['materiaux'][mat_name]['unite'] = unite_matiere
                
                # Traitements
                if prix_traitement_piece > 0 and traitement_code:
                    vol_pour_traitement = largeur_brute * hauteur_brute * longueur_utile / 1_000_000_000
                    details['traitements'][traitement_code]['qte'] += vol_pour_traitement
                    details['traitements'][traitement_code]['prix'] += prix_traitement_piece
                    details['traitements'][traitement_code]['unite'] = "m³"
                
                # Façonnages
                if prix_faconnage_piece > 0 and sku_faconnage:
                    if sku_faconnage.endswith("_L"):
                        qte_fac = longueur_utile_m
                        unite_fac = "ml"
                    else:
                        qte_fac = vol_utile if not mat_name.endswith("_L") else largeur_brute * hauteur_brute * longueur_utile / 1_000_000_000
                        unite_fac = "m³"
                    details['faconnages'][sku_faconnage]['qte'] += qte_fac
                    details['faconnages'][sku_faconnage]['prix'] += prix_faconnage_piece
                    details['faconnages'][sku_faconnage]['unite'] = unite_fac
                
                # Prestations
                if prix_prestation_piece > 0 and presta_code:
                    details['prestations'][presta_code]['qte'] += qte_presta
                    details['prestations'][presta_code]['prix'] += prix_prestation_piece
                    details['prestations'][presta_code]['unite'] = unite_presta

                # Accumulation par groupe - UTILISER LES PRIX DÉJÀ CALCULÉS
                val = ventilation_groupes[groupe_complet]
                val["nb_pieces"] += 1
                val["qte_achetee"] += qte_achetee
                val["qte_utile"] += qte_utile
                val["qte_chute_ventilee"] += qte_chute_piece
                val["matiere"] += prix_achat_piece  # Utiliser le prix achat calculé
                val["faconnage"] += prix_faconnage_piece
                val["traitement"] += prix_traitement_piece
                val["prestation"] += prix_prestation_piece
                val["surface_prestation"] += surface_presta
                val["volume_prestation"] += volume_presta
                val["chute_val"] += montant_chute_val

                self.elements_traites += 1
                
                # Mise à jour de la progression
                self.progress.config(value=i+1)
                self.update_stats()
                
            except Exception as e:
                self.log(f"❌ Erreur pour l'élément {eid} : {e}")
                
        self.groupes_detectes = len(ventilation_groupes)
        self.log(f"✅ Analyse terminée : {self.groupes_detectes} groupes détectés")
        self.log("💡 Utilisation des prix déjà calculés (pas de recalcul)")
        
        # Stocker les détails pour la présentation hiérarchique
        self.details_par_groupe = details_par_groupe
        
        return ventilation_groupes

    def generate_excel(self, ventilation_groupes):
        """Génère le fichier Excel avec les deux feuilles"""
        try:
            # Import openpyxl ici pour éviter les erreurs de chargement
            import openpyxl
            from openpyxl.worksheet.table import Table, TableStyleInfo
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment
            
            self.log("📊 Création du fichier Excel...")
            
            wb = openpyxl.Workbook()
            
            # === FEUILLE 1 : Analyse détaillée ===
            ws_analyse = wb.active
            ws_analyse.title = "Analyse détaillée"
            
            headers_analyse = ["Groupe", "Sous-groupe", "Nb Pièces", "Qté Achetée", "Qté Utile", "Qté Chute",
                             "Coût Matière (€)", "Coût Façonnage (€)", "Coût Traitement (€)", "Coût Prestation (€)",
                             "Surface Prestation (m²)", "Volume Prestation (m³)", "Valorisation Chute (€)", "TOTAL (€)"]
            ws_analyse.append(headers_analyse)

            for groupe_complet, val in ventilation_groupes.items():
                if " | " in groupe_complet:
                    groupe_principal, sous_groupe = groupe_complet.split(" | ", 1)
                else:
                    groupe_principal, sous_groupe = groupe_complet, "Non défini"
                    
                total_lot = val["matiere"] + val["faconnage"] + val["traitement"] + val["prestation"] + val["chute_val"]
                ws_analyse.append([
                    groupe_principal, sous_groupe,
                    int(val["nb_pieces"]), 
                    round(val["qte_achetee"], 4), 
                    round(val["qte_utile"], 4),
                    round(val["qte_chute_ventilee"], 4), 
                    round(val["matiere"], 2), 
                    round(val["faconnage"], 2),
                    round(val["traitement"], 2), 
                    round(val["prestation"], 2), 
                    round(val["surface_prestation"], 4),
                    round(val["volume_prestation"], 4), 
                    round(val["chute_val"], 2), 
                    round(total_lot, 2)
                ])

            # Total global pour feuille analyse
            total_row_analyse = ["TOTAL GLOBAL", ""] + ["" for _ in range(ws_analyse.max_column - 2)]
            for col in range(3, ws_analyse.max_column + 1):
                col_letter = get_column_letter(col)
                total_row_analyse[col - 1] = f"=SUBTOTAL(109,{col_letter}2:{col_letter}{ws_analyse.max_row})"
            ws_analyse.append(total_row_analyse)

            # Format tableau analyse
            tab_analyse = Table(displayName="TableAnalyseDetaillee", ref=f"A1:{get_column_letter(ws_analyse.max_column)}{ws_analyse.max_row}")
            tab_analyse.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
            ws_analyse.add_table(tab_analyse)

            # === FEUILLE 2 : Présentation hiérarchique ===
            ws_hierarchique = wb.create_sheet("Présentation devis")
            self.create_hierarchical_presentation(ws_hierarchique, ventilation_groupes)

            # Ajustement largeur colonnes pour les deux feuilles
            for ws in [ws_analyse, ws_hierarchique]:
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            # Nom dynamique fichier
            try:
                num_devis = uc.get_project_user_attribute(1).strip().replace(" ", "_")
                client = uc.get_project_number().strip().replace(" ", "_")
            except Exception:
                num_devis = "Devis"
                client = "Client"
                
            date_str = datetime.today().strftime("%Y-%m-%d")
            nom_fichier = f"{num_devis}-{client}-devis_par_groupe.xlsx"

            # Enregistrement fichier Excel
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            output_path = os.path.join(desktop, nom_fichier)
            wb.save(output_path)
            
            self.log(f"✅ Fichier Excel créé : {nom_fichier}")
            
            # Ouvrir le fichier
            subprocess.Popen(['start', '', output_path], shell=True)
            self.log("📂 Fichier ouvert automatiquement")

        except Exception as e:
            self.log(f"❌ Erreur génération Excel : {e}")
            raise

    def create_hierarchical_presentation(self, ws, ventilation_groupes):
        """Crée la présentation hiérarchique avec les vraies données extraites"""
        
        # Import des styles openpyxl ici
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Titre principal avec nom du projet
        try:
            num_devis = uc.get_project_user_attribute(1).strip()
            client = uc.get_project_number().strip()
            titre_devis = f"Devis {num_devis} - {client}"
        except Exception:
            titre_devis = "Devis"
            
        ws['A1'] = titre_devis
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.merge_cells('A1:F1')
        
        # Headers des colonnes
        ws['D2'] = "Quantité"
        ws['E2'] = "prix total"
        ws['D3'] = "m3-m2-ml"
        ws['E3'] = "€"
        
        # Style headers
        for cell in ['D2', 'E2', 'D3', 'E3']:
            ws[cell].fill = PatternFill(start_color="85C1E9", end_color="85C1E9", fill_type="solid")
            ws[cell].alignment = Alignment(horizontal="center")
            ws[cell].font = Font(bold=True)
        
        # Organiser les données par groupe principal
        groupes_organises = defaultdict(dict)
        for groupe_complet, data in ventilation_groupes.items():
            if " | " in groupe_complet:
                groupe_principal, sous_groupe = groupe_complet.split(" | ", 1)
            else:
                groupe_principal, sous_groupe = groupe_complet, "Non défini"
            
            if groupe_principal not in groupes_organises:
                groupes_organises[groupe_principal] = {}
            groupes_organises[groupe_principal][sous_groupe] = data
        
        current_row = 4
        
        for groupe_principal, sous_groupes in groupes_organises.items():
            # Ligne groupe principal avec nom
            ws[f'A{current_row}'] = f"Groupe: {groupe_principal}"
            ws[f'A{current_row}'].fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            # Calculer le total du groupe principal
            total_groupe_principal = 0
            
            for sous_groupe, data in sous_groupes.items():
                groupe_complet = f"{groupe_principal} | {sous_groupe}"
                details = self.details_par_groupe[groupe_complet]
                
                # Ligne sous-groupe avec nom
                ws[f'B{current_row}'] = f"sous groupe: {sous_groupe}"
                ws[f'B{current_row}'].fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
                ws[f'B{current_row}'].font = Font(bold=True)
                current_row += 1
                
                # Calculer le total du sous-groupe
                total_sous_groupe = 0
                
                # === MATIÈRE ===
                if details['materiaux']:
                    ws[f'C{current_row}'] = "Matière"
                    ws[f'C{current_row}'].fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
                    current_row += 1
                    
                    # Détails par matériau
                    for mat_name, mat_data in details['materiaux'].items():
                        if mat_data['prix'] > 0:
                            ws[f'C{current_row}'] = mat_name
                            ws[f'C{current_row}'].alignment = Alignment(horizontal="left")
                            ws[f'D{current_row}'] = round(mat_data['qte'], 3)
                            ws[f'E{current_row}'] = round(mat_data['prix'], 2)
                            total_sous_groupe += mat_data['prix']
                            current_row += 1
                
                # === PRESTATION ===
                if details['prestations']:
                    ws[f'C{current_row}'] = "Prestation"
                    ws[f'C{current_row}'].fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
                    current_row += 1
                    
                    # Détails par prestation
                    for presta_code, presta_data in details['prestations'].items():
                        if presta_data['prix'] > 0:
                            ws[f'C{current_row}'] = presta_code
                            ws[f'C{current_row}'].alignment = Alignment(horizontal="left")
                            ws[f'D{current_row}'] = round(presta_data['qte'], 3)
                            ws[f'E{current_row}'] = round(presta_data['prix'], 2)
                            total_sous_groupe += presta_data['prix']
                            current_row += 1
                
                # === TRAITEMENT ===
                if details['traitements']:
                    ws[f'C{current_row}'] = "Traitement"
                    ws[f'C{current_row}'].fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
                    current_row += 1
                    
                    # Détails par traitement
                    for trait_code, trait_data in details['traitements'].items():
                        if trait_data['prix'] > 0:
                            ws[f'C{current_row}'] = trait_code
                            ws[f'C{current_row}'].alignment = Alignment(horizontal="left")
                            ws[f'D{current_row}'] = round(trait_data['qte'], 3)
                            ws[f'E{current_row}'] = round(trait_data['prix'], 2)
                            total_sous_groupe += trait_data['prix']
                            current_row += 1
                
                # === FAÇONNAGE (CORRIGÉ) ===
                if details['faconnages'] or data["chute_val"] > 0:
                    ws[f'C{current_row}'] = "Façonnage"
                    ws[f'C{current_row}'].fill = PatternFill(start_color="F1C40F", end_color="F1C40F", fill_type="solid")
                    current_row += 1
                    
                    # Détails par façonnage
                    for fac_code, fac_data in details['faconnages'].items():
                        if fac_data['prix'] > 0:
                            ws[f'C{current_row}'] = fac_code
                            ws[f'C{current_row}'].alignment = Alignment(horizontal="left")
                            ws[f'D{current_row}'] = round(fac_data['qte'], 3)
                            ws[f'E{current_row}'] = round(fac_data['prix'], 2)
                            total_sous_groupe += fac_data['prix']
                            current_row += 1
                    
                    # AJOUT : Valorisation des chutes dans façonnage
                    if data["chute_val"] > 0:
                        ws[f'C{current_row}'] = "Valorisation chutes"
                        ws[f'C{current_row}'].alignment = Alignment(horizontal="left")
                        ws[f'D{current_row}'] = round(data["qte_chute_ventilee"], 3)
                        ws[f'E{current_row}'] = round(data["chute_val"], 2)
                        total_sous_groupe += data["chute_val"]
                        current_row += 1
                
                # Total sous-groupe
                ws[f'C{current_row}'] = "TOTAL sous-groupe"
                ws[f'E{current_row}'] = round(total_sous_groupe, 2)
                ws[f'C{current_row}'].font = Font(bold=True)
                ws[f'E{current_row}'].font = Font(bold=True)
                ws[f'C{current_row}'].fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
                ws[f'E{current_row}'].fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
                current_row += 2  # Espacement
                
                total_groupe_principal += total_sous_groupe
            
            # Total groupe principal
            ws[f'B{current_row}'] = "TOTAL GROUPE"
            ws[f'E{current_row}'] = round(total_groupe_principal, 2)
            ws[f'B{current_row}'].font = Font(bold=True, size=12)
            ws[f'E{current_row}'].font = Font(bold=True, size=12)
            ws[f'B{current_row}'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            ws[f'E{current_row}'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            current_row += 3  # Espacement entre groupes

def main():
    """Fonction principale"""
    app = DevisGroupeInterface()
    app.root.mainloop()

if __name__ == "__main__":
    main()