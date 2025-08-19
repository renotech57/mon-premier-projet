import sys
import os
sys.path.append(r"C:\cadwork\libs")

import cadwork
import geometry_controller as gc
import attribute_controller as ac
import element_controller as ec
import material_controller as mc
import utility_controller as uc
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading

class ConfigurateurPrixInterface:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Configurateur Prix par Pièce - Cadwork")
        self.root.geometry("1200x800")
        self.root.configure(bg='#f0f0f0')
        
        # Constantes
        self.prix_traitement = {"CL2": 25.5, "CL3": 215, "CL2 INCOLORE": 50}
        self.prix_faconnage = {
            "T1_V": 224.4, "T16_L": 2.34, "T2_V": 243.1, "T3_V": 320, "T4_V": 263.5
        }
        self.prix_prestation = {
            "MO1_S": 17, "MO2_S": 28.9, "MO3_S": 22.95, "MO4_S": 19.55,
            "MO5_V": 219.3, "MO6_V": 85, "MO7_V": 119, "MO8_V": 35
        }
        self.VALORISATION_CHUTE = 80
        
        # Variables
        self.element_ids = []
        self.materiaux_detectes = {}
        self.config_materiaux = {}
        self.should_stop = False
        self.config_validee = False
        self.liste_pieces = []
        
        self.setup_ui()
        
    def setup_ui(self):
        # Style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Titre principal
        title_frame = tk.Frame(self.root, bg='#2c3e50', height=80)
        title_frame.pack(fill='x', padx=0, pady=0)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="⚙️ CONFIGURATEUR PRIX PAR PIÈCE", 
                              font=('Arial', 18, 'bold'), fg='white', bg='#2c3e50')
        title_label.pack(expand=True)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Zone d'information
        info_frame = tk.LabelFrame(main_frame, text="ℹ️ Instructions", font=('Arial', 12, 'bold'), 
                                 bg='#f0f0f0', fg='#2c3e50')
        info_frame.pack(fill='x', pady=(0, 10))
        
        info_text = tk.Label(info_frame, 
                           text="1. Cliquez 'ANALYSER' pour détecter tous les matériaux\n"
                                "2. Configurez chaque matériau selon vos besoins (optimisation, calculs, unités)\n"
                                "3. Validez la configuration\n"
                                "4. Lancez le calcul des prix avec les paramètres choisis",
                           font=('Arial', 10), bg='#f0f0f0', fg='#2c3e50', justify='left')
        info_text.pack(pady=10, padx=10)
        
        # Boutons d'action principaux
        action_frame = tk.Frame(main_frame, bg='#f0f0f0')
        action_frame.pack(fill='x', pady=(0, 10))
        
        self.analyze_button = tk.Button(action_frame, text="🔍 ANALYSER LES MATÉRIAUX", 
                                      command=self.analyze_materials, font=('Arial', 12, 'bold'),
                                      bg='#3498db', fg='white', height=2, cursor='hand2')
        self.analyze_button.pack(side='left', padx=(0, 10))
        
        self.validate_button = tk.Button(action_frame, text="✅ VALIDER CONFIGURATION", 
                                       command=self.validate_config, font=('Arial', 12, 'bold'),
                                       bg='#27ae60', fg='white', height=2, cursor='hand2',
                                       state='disabled')
        self.validate_button.pack(side='left', padx=(0, 10))
        
        self.calculate_button = tk.Button(action_frame, text="🧮 CALCULER PRIX", 
                                        command=self.calculate_prices, font=('Arial', 12, 'bold'),
                                        bg='#e67e22', fg='white', height=2, cursor='hand2',
                                        state='disabled')
        self.calculate_button.pack(side='left', padx=(0, 10))
        
        # Zone de configuration des matériaux
        config_frame = tk.LabelFrame(main_frame, text="⚙️ Configuration des matériaux", 
                                   font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        config_frame.pack(fill='both', expand=True, pady=(0, 10))
        
        # Frame avec scrollbar pour la configuration
        canvas_frame = tk.Frame(config_frame)
        canvas_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.canvas = tk.Canvas(canvas_frame, bg='white')
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='white')
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Zone de log
        log_frame = tk.LabelFrame(main_frame, text="📋 Journal", font=('Arial', 12, 'bold'), 
                                bg='#f0f0f0', fg='#2c3e50')
        log_frame.pack(fill='x', pady=(10, 0))
        
        self.text_area = scrolledtext.ScrolledText(log_frame, height=8, width=80, 
                                                 font=('Courier', 9), bg='#ffffff', 
                                                 fg='#2c3e50', wrap=tk.WORD)
        self.text_area.pack(fill='x', padx=10, pady=10)
        
        # Bouton fermer
        close_frame = tk.Frame(main_frame, bg='#f0f0f0')
        close_frame.pack(fill='x', pady=(10, 0))
        
        self.close_button = tk.Button(close_frame, text="🚪 FERMER", 
                                    command=self.close_application, font=('Arial', 12, 'bold'),
                                    bg='#95a5a6', fg='white', height=2, cursor='hand2')
        self.close_button.pack(side='right')
        
        # Gestionnaire de fermeture de fenêtre
        self.root.protocol("WM_DELETE_WINDOW", self.close_application)
        
    def close_application(self):
        """Ferme l'application proprement"""
        self.root.destroy()
        
    def log(self, message):
        """Ajoute un message dans la zone de texte"""
        self.text_area.config(state='normal')
        self.text_area.insert(tk.END, f"{message}\n")
        self.text_area.config(state='disabled')
        self.text_area.see(tk.END)
        self.root.update()
        
    def analyze_materials(self):
        """Analyse les éléments et détecte tous les matériaux"""
        try:
            self.log("🔍 Récupération des éléments...")
            self.element_ids = ec.get_active_identifiable_element_ids()
            
            if not self.element_ids:
                messagebox.showwarning("Aucun élément", "Veuillez sélectionner des éléments avant l'analyse.")
                return
                
            self.log(f"📊 Analyse de {len(self.element_ids)} élément(s)")
            
            # Détection des matériaux
            materiaux_stats = defaultdict(lambda: {'count': 0, 'exemple_id': None})
            
            for eid in self.element_ids:
                try:
                    mat_name = ac.get_element_material_name(eid).strip()
                    materiaux_stats[mat_name]['count'] += 1
                    if materiaux_stats[mat_name]['exemple_id'] is None:
                        materiaux_stats[mat_name]['exemple_id'] = eid
                except Exception:
                    continue
            
            self.materiaux_detectes = dict(materiaux_stats)
            self.log(f"✅ {len(self.materiaux_detectes)} matériau(x) détecté(s)")
            
            # Créer l'interface de configuration
            self.create_config_interface()
            
            # Activer le bouton de validation
            self.validate_button.config(state='normal')
            
        except Exception as e:
            self.log(f"❌ Erreur lors de l'analyse : {e}")
            messagebox.showerror("Erreur", f"Erreur lors de l'analyse :\n{e}")
    
    def create_config_interface(self):
        """Crée l'interface de configuration pour chaque matériau"""
        # Nettoyer le frame existant
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        self.config_vars = {}
        
        # Headers
        header_frame = tk.Frame(self.scrollable_frame, bg='white')
        header_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(header_frame, text="Matériau", font=('Arial', 10, 'bold'), 
                bg='white', width=22, anchor='w').grid(row=0, column=0, padx=3)
        tk.Label(header_frame, text="Qté", font=('Arial', 10, 'bold'), 
                bg='white', width=4).grid(row=0, column=1, padx=3)
        tk.Label(header_frame, text="Optimiser", font=('Arial', 10, 'bold'), 
                bg='white', width=8).grid(row=0, column=2, padx=3)
        tk.Label(header_frame, text="Matière", font=('Arial', 10, 'bold'), 
                bg='white', width=8).grid(row=0, column=3, padx=3)
        tk.Label(header_frame, text="Chutes", font=('Arial', 10, 'bold'), 
                bg='white', width=6).grid(row=0, column=4, padx=3)
        tk.Label(header_frame, text="Unité", font=('Arial', 10, 'bold'), 
                bg='white', width=6).grid(row=0, column=5, padx=3)
        tk.Label(header_frame, text="Façonnage", font=('Arial', 10, 'bold'), 
                bg='white', width=9).grid(row=0, column=6, padx=3)
        tk.Label(header_frame, text="Traitement", font=('Arial', 10, 'bold'), 
                bg='white', width=9).grid(row=0, column=7, padx=3)
        tk.Label(header_frame, text="Prestation", font=('Arial', 10, 'bold'), 
                bg='white', width=9).grid(row=0, column=8, padx=3)
        
        # Ligne de séparation
        separator = tk.Frame(self.scrollable_frame, height=2, bg='#bdc3c7')
        separator.pack(fill='x', padx=5, pady=2)
        
        # Configuration pour chaque matériau
        for mat_name, stats in self.materiaux_detectes.items():
            self.config_vars[mat_name] = {}
            
            mat_frame = tk.Frame(self.scrollable_frame, bg='white', relief='ridge', bd=1)
            mat_frame.pack(fill='x', padx=5, pady=2)
            
            # Nom du matériau
            tk.Label(mat_frame, text=mat_name, font=('Arial', 9), 
                    bg='white', width=22, anchor='w').grid(row=0, column=0, padx=3, pady=3)
            
            # Quantité
            tk.Label(mat_frame, text=str(stats['count']), font=('Arial', 9), 
                    bg='white', width=4).grid(row=0, column=1, padx=3, pady=3)
            
            # À optimiser
            self.config_vars[mat_name]['optimiser'] = tk.BooleanVar()
            tk.Checkbutton(mat_frame, variable=self.config_vars[mat_name]['optimiser'],
                          bg='white').grid(row=0, column=2, padx=3, pady=3)
            
            # Compter en matière
            self.config_vars[mat_name]['matiere'] = tk.BooleanVar()
            tk.Checkbutton(mat_frame, variable=self.config_vars[mat_name]['matiere'],
                          bg='white').grid(row=0, column=3, padx=3, pady=3)
            
            # Calculer les chutes
            self.config_vars[mat_name]['chutes'] = tk.BooleanVar()
            tk.Checkbutton(mat_frame, variable=self.config_vars[mat_name]['chutes'],
                          bg='white').grid(row=0, column=4, padx=3, pady=3)
            
            # Unité
            self.config_vars[mat_name]['unite'] = tk.StringVar(value="m³")
            unite_combo = ttk.Combobox(mat_frame, textvariable=self.config_vars[mat_name]['unite'],
                                     values=["m³", "m²", "ml"], width=5, state="readonly")
            unite_combo.grid(row=0, column=5, padx=3, pady=3)
            
            # Façonnage autorisé
            self.config_vars[mat_name]['faconnage'] = tk.BooleanVar(value=True)
            tk.Checkbutton(mat_frame, variable=self.config_vars[mat_name]['faconnage'],
                          bg='white').grid(row=0, column=6, padx=3, pady=3)
            
            # Traitement autorisé
            self.config_vars[mat_name]['traitement'] = tk.BooleanVar(value=True)
            tk.Checkbutton(mat_frame, variable=self.config_vars[mat_name]['traitement'],
                          bg='white').grid(row=0, column=7, padx=3, pady=3)
            
            # Prestations autorisées
            self.config_vars[mat_name]['prestation'] = tk.BooleanVar(value=True)
            tk.Checkbutton(mat_frame, variable=self.config_vars[mat_name]['prestation'],
                          bg='white').grid(row=0, column=8, padx=3, pady=3)
            
            # Configuration par défaut basée sur le nom du matériau
            self.set_default_config(mat_name)
        
        # Boutons de configuration rapide
        quick_frame = tk.Frame(self.scrollable_frame, bg='white')
        quick_frame.pack(fill='x', padx=5, pady=10)
        
        tk.Label(quick_frame, text="Configuration rapide :", font=('Arial', 10, 'bold'), 
                bg='white').pack(side='left', padx=5)
        
        tk.Button(quick_frame, text="Bois structure", command=self.config_bois_structure,
                 bg='#27ae60', fg='white', font=('Arial', 9)).pack(side='left', padx=2)
        tk.Button(quick_frame, text="Panneaux", command=self.config_panneaux,
                 bg='#f39c12', fg='white', font=('Arial', 9)).pack(side='left', padx=2)
        tk.Button(quick_frame, text="Enveloppes", command=self.config_enveloppes,
                 bg='#e74c3c', fg='white', font=('Arial', 9)).pack(side='left', padx=2)
        tk.Button(quick_frame, text="Linéaires", command=self.config_lineaires,
                 bg='#9b59b6', fg='white', font=('Arial', 9)).pack(side='left', padx=2)
        tk.Button(quick_frame, text="Reset", command=self.config_reset,
                 bg='#95a5a6', fg='white', font=('Arial', 9)).pack(side='left', padx=2)
    
    def set_default_config(self, mat_name):
        """Définit une configuration par défaut basée sur le nom du matériau"""
        mat_lower = mat_name.lower()
        
        if any(x in mat_lower for x in ['enveloppe', 'paroi', 'isolation']):
            # Enveloppes : que prestations
            self.config_vars[mat_name]['optimiser'].set(False)
            self.config_vars[mat_name]['matiere'].set(False)
            self.config_vars[mat_name]['chutes'].set(False)
            self.config_vars[mat_name]['unite'].set("m²")
            self.config_vars[mat_name]['faconnage'].set(False)
            self.config_vars[mat_name]['traitement'].set(False)
            self.config_vars[mat_name]['prestation'].set(True)
        elif any(x in mat_lower for x in ['osb', 'panneau', 'contreplaque', 'mdf']):
            # Panneaux : matière au m² sans optimisation
            self.config_vars[mat_name]['optimiser'].set(False)
            self.config_vars[mat_name]['matiere'].set(True)
            self.config_vars[mat_name]['chutes'].set(False)
            self.config_vars[mat_name]['unite'].set("m²")
            self.config_vars[mat_name]['faconnage'].set(True)
            self.config_vars[mat_name]['traitement'].set(True)
            self.config_vars[mat_name]['prestation'].set(True)
        elif any(x in mat_lower for x in ['_l', 'barre', 'profilé']):
            # Matériaux linéaires : optimisation + matière au ml
            self.config_vars[mat_name]['optimiser'].set(True)
            self.config_vars[mat_name]['matiere'].set(True)
            self.config_vars[mat_name]['chutes'].set(False)  # Pas de valorisation pour linéaires
            self.config_vars[mat_name]['unite'].set("ml")
            self.config_vars[mat_name]['faconnage'].set(True)
            self.config_vars[mat_name]['traitement'].set(True)
            self.config_vars[mat_name]['prestation'].set(True)
        else:
            # Bois par défaut : optimisation + matière + chutes au m³
            self.config_vars[mat_name]['optimiser'].set(True)
            self.config_vars[mat_name]['matiere'].set(True)
            self.config_vars[mat_name]['chutes'].set(True)
            self.config_vars[mat_name]['unite'].set("m³")
            self.config_vars[mat_name]['faconnage'].set(True)
            self.config_vars[mat_name]['traitement'].set(True)
            self.config_vars[mat_name]['prestation'].set(True)
    
    def config_bois_structure(self):
        """Configuration rapide pour bois de structure"""
        for mat_name in self.materiaux_detectes.keys():
            mat_lower = mat_name.lower()
            if any(x in mat_lower for x in ['epicea', 'sapin', 'douglas', 'pin', 'meleze', 'chene', 'kvh', 'gl24', 'bois']):
                self.config_vars[mat_name]['optimiser'].set(True)
                self.config_vars[mat_name]['matiere'].set(True)
                self.config_vars[mat_name]['chutes'].set(True)
                self.config_vars[mat_name]['unite'].set("m³")
                self.config_vars[mat_name]['faconnage'].set(True)
                self.config_vars[mat_name]['traitement'].set(True)
                self.config_vars[mat_name]['prestation'].set(True)
    
    def config_panneaux(self):
        """Configuration rapide pour panneaux"""
        for mat_name in self.materiaux_detectes.keys():
            mat_lower = mat_name.lower()
            if any(x in mat_lower for x in ['osb', 'panneau', 'contreplaque', 'mdf', 'agglomere']):
                self.config_vars[mat_name]['optimiser'].set(False)
                self.config_vars[mat_name]['matiere'].set(True)
                self.config_vars[mat_name]['chutes'].set(False)
                self.config_vars[mat_name]['unite'].set("m²")
                self.config_vars[mat_name]['faconnage'].set(True)
                self.config_vars[mat_name]['traitement'].set(True)
                self.config_vars[mat_name]['prestation'].set(True)
    
    def config_enveloppes(self):
        """Configuration rapide pour enveloppes"""
        for mat_name in self.materiaux_detectes.keys():
            mat_lower = mat_name.lower()
            if any(x in mat_lower for x in ['enveloppe', 'paroi', 'isolation', 'bardage']):
                self.config_vars[mat_name]['optimiser'].set(False)
                self.config_vars[mat_name]['matiere'].set(False)
                self.config_vars[mat_name]['chutes'].set(False)
                self.config_vars[mat_name]['unite'].set("m²")
                self.config_vars[mat_name]['faconnage'].set(False)
                self.config_vars[mat_name]['traitement'].set(False)
                self.config_vars[mat_name]['prestation'].set(True)
    
    def config_lineaires(self):
        """Configuration rapide pour matériaux linéaires"""
        for mat_name in self.materiaux_detectes.keys():
            mat_lower = mat_name.lower()
            if any(x in mat_lower for x in ['_l', 'barre', 'profilé', 'tube', 'rail']):
                self.config_vars[mat_name]['optimiser'].set(True)
                self.config_vars[mat_name]['matiere'].set(True)
                self.config_vars[mat_name]['chutes'].set(False)
                self.config_vars[mat_name]['unite'].set("ml")
                self.config_vars[mat_name]['faconnage'].set(True)
                self.config_vars[mat_name]['traitement'].set(True)
                self.config_vars[mat_name]['prestation'].set(True)
    
    def config_reset(self):
        """Remet tout à zéro"""
        for mat_name in self.materiaux_detectes.keys():
            self.config_vars[mat_name]['optimiser'].set(False)
            self.config_vars[mat_name]['matiere'].set(False)
            self.config_vars[mat_name]['chutes'].set(False)
            self.config_vars[mat_name]['unite'].set("m³")
            self.config_vars[mat_name]['faconnage'].set(False)
            self.config_vars[mat_name]['traitement'].set(False)
            self.config_vars[mat_name]['prestation'].set(False)
    
    def validate_config(self):
        """Valide la configuration et prépare les paramètres"""
        try:
            self.config_materiaux = {}
            
            for mat_name in self.materiaux_detectes.keys():
                self.config_materiaux[mat_name] = {
                    'optimiser': self.config_vars[mat_name]['optimiser'].get(),
                    'matiere': self.config_vars[mat_name]['matiere'].get(),
                    'chutes': self.config_vars[mat_name]['chutes'].get(),
                    'unite': self.config_vars[mat_name]['unite'].get(),
                    'faconnage': self.config_vars[mat_name]['faconnage'].get(),
                    'traitement': self.config_vars[mat_name]['traitement'].get(),
                    'prestation': self.config_vars[mat_name]['prestation'].get()
                }
            
            self.config_validee = True
            self.calculate_button.config(state='normal')
            
            # Résumé de la configuration
            self.log("✅ Configuration validée !")
            self.log("=" * 50)
            
            count_optimiser = sum(1 for c in self.config_materiaux.values() if c['optimiser'])
            count_matiere = sum(1 for c in self.config_materiaux.values() if c['matiere'])
            count_chutes = sum(1 for c in self.config_materiaux.values() if c['chutes'])
            count_faconnage = sum(1 for c in self.config_materiaux.values() if c['faconnage'])
            count_traitement = sum(1 for c in self.config_materiaux.values() if c['traitement'])
            count_prestation = sum(1 for c in self.config_materiaux.values() if c['prestation'])
            
            self.log(f"🔧 {count_optimiser} matériau(x) à optimiser")
            self.log(f"📦 {count_matiere} matériau(x) comptés en matière")
            self.log(f"♻️ {count_chutes} matériau(x) avec valorisation chutes")
            self.log(f"⚙️ {count_faconnage} matériau(x) avec façonnage")
            self.log(f"🎨 {count_traitement} matériau(x) avec traitement")
            self.log(f"🛠️ {count_prestation} matériau(x) avec prestation")
            self.log("=" * 50)
            
        except Exception as e:
            self.log(f"❌ Erreur validation : {e}")
            messagebox.showerror("Erreur", f"Erreur lors de la validation :\n{e}")
    
    def calculate_prices(self):
        """Lance le calcul des prix avec la configuration validée"""
        if not self.config_validee:
            messagebox.showwarning("Configuration requise", "Veuillez d'abord valider la configuration.")
            return
        
        try:
            self.log("🚀 Calcul des prix en cours...")
            
            # Vérification de l'optimisation pour les matériaux configurés
            elements_non_optimises = self.verifier_optimisation()
            
            if elements_non_optimises:
                self.log(f"❌ {len(elements_non_optimises)} élément(s) nécessitent une optimisation")
                for elem in elements_non_optimises:
                    self.log(f"   - {elem['materiau']} (ID: {elem['id']})")
                messagebox.showerror("Optimisation requise", 
                                   f"{len(elements_non_optimises)} élément(s) doivent être optimisés.\n"
                                   "Consultez le journal pour plus de détails.")
                return
            
            # Calcul des prix
            self.log("💰 Calcul des prix par pièce...")
            self.calculate_all_prices()
            
            self.log("📊 Génération du fichier Excel...")
            self.export_to_excel()
            
            self.log("🎯 Calcul terminé avec succès !")
            messagebox.showinfo("Succès", f"Calcul terminé !\n{len(self.liste_pieces)} pièces traitées")
            
        except Exception as e:
            self.log(f"❌ Erreur calcul : {e}")
            messagebox.showerror("Erreur", f"Erreur lors du calcul :\n{e}")
    
    def verifier_optimisation(self):
        """Vérifie l'optimisation selon la configuration"""
        elements_non_optimises = []
        
        for eid in self.element_ids:
            try:
                mat_name = ac.get_element_material_name(eid).strip()
                config = self.config_materiaux.get(mat_name, {})
                
                # Si le matériau doit être optimisé
                if config.get('optimiser', False):
                    num_barre = ac.get_user_attribute(eid, 12)
                    taux_chute = ac.get_user_attribute(eid, 13)
                    
                    num_barre_vide = not num_barre or str(num_barre).strip() == ""
                    taux_chute_vide = not taux_chute or str(taux_chute).strip() == ""
                    
                    if num_barre_vide or taux_chute_vide:
                        elements_non_optimises.append({
                            'id': eid,
                            'materiau': mat_name
                        })
                        
            except Exception:
                continue
                
        return elements_non_optimises
    
    def safe_float(self, val, default=0.0):
        try:
            return float(str(val).replace('%', '').replace(',', '.'))
        except (ValueError, TypeError):
            return default
    
    def calculate_all_prices(self):
        """Calcule les prix pour toutes les pièces selon la configuration"""
        # Dictionnaires de recap
        recap_traitement = defaultdict(float)
        recap_qte_traitement = defaultdict(float)
        recap_faconnage = defaultdict(float)
        recap_qte_faconnage = defaultdict(float)
        recap_prestation = defaultdict(float)
        recap_qte_prestation = defaultdict(float)
        recap_materiaux_piece = defaultdict(float)
        recap_qte_materiaux_piece = defaultdict(float)
        recap_materiaux_achat = defaultdict(float)
        recap_qte_materiaux_achat = defaultdict(float)
        recap_prix_unitaire_materiaux = {}
        self.liste_pieces = []
        
        total_volume_chute = 0
        total_surface_chute = 0
        total_longueur_chute = 0
        
        for i, eid in enumerate(self.element_ids):
            try:
                # Configuration du matériau
                mat_name = ac.get_element_material_name(eid).strip()
                config = self.config_materiaux.get(mat_name, {})
                
                # Données géométriques
                largeur = round(gc.get_list_width(eid))
                hauteur = round(gc.get_list_height(eid))
                longueur_mm = gc.get_length(eid)
                longueur_m = longueur_mm / 1000
                section_str = f"{largeur}x{hauteur}"
                
                # Prix unitaire matière
                mat_id = mc.get_material_id(mat_name)
                prix_u = self.safe_float(mc.get_price(mat_id))
                recap_prix_unitaire_materiaux[mat_name] = prix_u
                
                # Taux de chute selon configuration
                if config.get('optimiser', False):
                    taux_chute = self.safe_float(ac.get_user_attribute(eid, 13)) / 100.0
                    num_barre = ac.get_user_attribute(eid, 12) or "N/A"
                else:
                    taux_chute = 0.0
                    num_barre = "N/A"
                
                # Calculs selon l'unité configurée
                unite = config.get('unite', 'm³')
                
                if unite == 'ml':
                    qte_utile = longueur_m
                    qte_achetee = qte_utile / (1.0 - taux_chute) if taux_chute < 1.0 else qte_utile
                    qte_chute_piece = qte_achetee - qte_utile
                    total_longueur_chute += qte_chute_piece
                elif unite == 'm²':
                    surface_m2 = (largeur * longueur_mm) / 1_000_000
                    qte_utile = surface_m2
                    qte_achetee = qte_utile / (1.0 - taux_chute) if taux_chute < 1.0 else qte_utile
                    qte_chute_piece = qte_achetee - qte_utile
                    total_surface_chute += qte_chute_piece
                else:  # m³
                    volume_m3 = largeur * hauteur * longueur_mm / 1_000_000_000
                    qte_utile = volume_m3
                    qte_achetee = qte_utile / (1.0 - taux_chute) if taux_chute < 1.0 else qte_utile
                    qte_chute_piece = qte_achetee - qte_utile
                    total_volume_chute += qte_chute_piece
                
                # Prix matière selon configuration
                if config.get('matiere', False):
                    prix_piece = round(qte_utile * prix_u, 2)
                    prix_achat = round(qte_achetee * prix_u, 2)
                    recap_materiaux_piece[mat_name] += prix_piece
                    recap_qte_materiaux_piece[mat_name] += qte_utile
                    recap_materiaux_achat[mat_name] += prix_achat
                    recap_qte_materiaux_achat[mat_name] += qte_achetee
                else:
                    prix_piece = 0.0
                    prix_achat = 0.0
                
                # Traitement selon configuration
                prix_trait_piece = 0.0
                if config.get('traitement', False):
                    traitement_code = ac.get_user_attribute(eid, 1)
                    if traitement_code and traitement_code.strip() in self.prix_traitement:
                        tarif = self.prix_traitement[traitement_code.strip()]
                        # Volume pour traitement (toujours en m³)
                        vol_traitement = largeur * hauteur * longueur_mm / 1_000_000_000
                        prix_trait_piece = round(vol_traitement * tarif, 2)
                        recap_qte_traitement[traitement_code.strip()] += vol_traitement
                
                # Façonnage selon configuration
                prix_fac = 0.0
                if config.get('faconnage', False):
                    sku = ac.get_sku(eid)
                    if sku and sku.strip() in self.prix_faconnage:
                        prix_f = self.prix_faconnage[sku.strip()]
                        if sku.strip().endswith("_L"):
                            qty_fac = longueur_m
                        else:
                            qty_fac = largeur * hauteur * longueur_mm / 1_000_000_000
                        prix_fac = round(qty_fac * prix_f, 2)
                        recap_faconnage[sku.strip()] += prix_fac
                        recap_qte_faconnage[sku.strip()] += qty_fac
                
                # Prestation selon configuration
                prix_presta = 0.0
                if config.get('prestation', False):
                    presta_code = ac.get_user_attribute(eid, 2)
                    if presta_code and presta_code.strip() in self.prix_prestation:
                        prix_p = self.prix_prestation[presta_code.strip()]
                        if presta_code.strip().endswith("_S"):
                            surface_presta = gc.get_element_reference_face_area(eid) / 1_000_000
                            qty_p = surface_presta
                        else:
                            qty_p = largeur * hauteur * longueur_mm / 1_000_000_000
                        prix_presta = round(qty_p * prix_p, 2)
                        recap_prestation[presta_code.strip()] += prix_presta
                        recap_qte_prestation[presta_code.strip()] += qty_p
                
                # Total pièce
                total_piece = round(prix_achat + prix_fac + prix_presta + prix_trait_piece, 2)
                
                # Mise à jour des attributs si prix > 0
                if total_piece > 0:
                    ac.set_user_attribute([eid], 8, str(prix_piece))
                    ac.set_user_attribute([eid], 9, str(prix_fac))
                    ac.set_user_attribute([eid], 10, str(prix_presta))
                    ac.set_user_attribute([eid], 11, str(total_piece))
                    ac.set_user_attribute([eid], 15, str(prix_achat))
                    ac.set_user_attribute([eid], 7, str(prix_trait_piece))
                
                # Liste pour Excel
                self.liste_pieces.append([
                    eid, mat_name, section_str, round(longueur_m, 3),
                    round(qte_utile, 5), round(qte_achetee, 5), round(prix_u, 2),
                    prix_piece, prix_achat, prix_trait_piece, prix_fac, prix_presta, total_piece,
                    num_barre, round(taux_chute * 100, 2), unite
                ])
                
            except Exception as e:
                self.log(f"❌ Erreur élément {eid}: {e}")
        
        # Valorisation des chutes selon configuration
        valorisation_chute = 0.0
        for mat_name, config in self.config_materiaux.items():
            if config.get('chutes', False):
                unite = config.get('unite', 'm³')
                if unite == 'm²':
                    valorisation_chute += total_surface_chute * self.VALORISATION_CHUTE
                elif unite == 'm³':
                    valorisation_chute += total_volume_chute * self.VALORISATION_CHUTE
                # Pas de valorisation pour ml
        
        # Récap traitement : calcul du total à la fin
        for code in recap_qte_traitement:
            recap_traitement[code] = round(recap_qte_traitement[code] * self.prix_traitement[code], 2)
        
        self.log(f"✅ {len(self.liste_pieces)} pièces calculées")
        self.log(f"💰 Valorisation des chutes : {round(valorisation_chute, 2)}€")
        
        # Stocker les recaps pour l'export
        self.recap_data = {
            'recap_materiaux_piece': recap_materiaux_piece,
            'recap_qte_materiaux_piece': recap_qte_materiaux_piece,
            'recap_materiaux_achat': recap_materiaux_achat,
            'recap_qte_materiaux_achat': recap_qte_materiaux_achat,
            'recap_traitement': recap_traitement,
            'recap_qte_traitement': recap_qte_traitement,
            'recap_faconnage': recap_faconnage,
            'recap_qte_faconnage': recap_qte_faconnage,
            'recap_prestation': recap_prestation,
            'recap_qte_prestation': recap_qte_prestation,
            'recap_prix_unitaire_materiaux': recap_prix_unitaire_materiaux,
            'valorisation_chute': round(valorisation_chute, 2),
            'total_volume_chute': total_volume_chute,
            'total_surface_chute': total_surface_chute,
            'total_longueur_chute': total_longueur_chute
        }
    
    def export_to_excel(self):
        """Exporte les résultats vers Excel"""
        try:
            wb = openpyxl.Workbook()
            ws_pieces = wb.active
            ws_pieces.title = "Pièces détaillées"
            ws_recap = wb.create_sheet("Récapitulatif")

            headers = [
                "ID Pièce", "Matériau", "Section (mm)", "Longueur (m)", "Quantité utile",
                "Quantité achetée", "Prix unitaire (€)", "Prix matière (€)", "Prix achat (€)", "Prix traitement (€)",
                "Prix façonnage (€)", "Prix prestation (€)", "Total pièce (€)", "N° Barre", "Taux chute (%)", "Unité"
            ]
            ws_pieces.append(headers)
            for row in self.liste_pieces:
                ws_pieces.append(row)

            tab1 = Table(displayName="TablePieces", ref=f"A1:{get_column_letter(ws_pieces.max_column)}{ws_pieces.max_row}")
            style1 = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab1.tableStyleInfo = style1
            ws_pieces.add_table(tab1)

            recap_headers = ["Catégorie", "Code", "Montant (€)", "Quantité", "Unité", "Prix unitaire (€)"]
            ws_recap.append(recap_headers)
            
            # Récap données avec unités configurées
            for mat, val in self.recap_data['recap_materiaux_piece'].items():
                config = self.config_materiaux.get(mat, {})
                unite = config.get('unite', 'm³')
                ws_recap.append(["Matière utile", mat, round(val,2), round(self.recap_data['recap_qte_materiaux_piece'][mat],3), unite, round(self.recap_data['recap_prix_unitaire_materiaux'][mat],2)])
            
            for mat, val in self.recap_data['recap_materiaux_achat'].items():
                config = self.config_materiaux.get(mat, {})
                unite = config.get('unite', 'm³')
                ws_recap.append(["Matière achat", mat, round(val,2), round(self.recap_data['recap_qte_materiaux_achat'][mat],3), unite, round(self.recap_data['recap_prix_unitaire_materiaux'][mat],2)])
            
            for code, val in self.recap_data['recap_traitement'].items():
                ws_recap.append(["Traitement", code, round(val,2), round(self.recap_data['recap_qte_traitement'][code],3), "m³", round(self.prix_traitement[code],2)])
            
            for code, val in self.recap_data['recap_faconnage'].items():
                unite_fac = "ml" if code.endswith("_L") else "m³"
                ws_recap.append(["Façonnage", code, round(val,2), round(self.recap_data['recap_qte_faconnage'][code],3), unite_fac, round(self.prix_faconnage[code],2)])
            
            for code, val in self.recap_data['recap_prestation'].items():
                unite_prest = "m²" if code.endswith("_S") else "m³"
                ws_recap.append(["Prestation", code, round(val,2), round(self.recap_data['recap_qte_prestation'][code],3), unite_prest, round(self.prix_prestation[code],2)])
            
            if self.recap_data['valorisation_chute'] > 0:
                ws_recap.append([
                    "Valorisation chute", "", self.recap_data['valorisation_chute'],
                    round(self.recap_data['total_volume_chute'] + self.recap_data['total_surface_chute'], 3), 
                    "m³/m²", self.VALORISATION_CHUTE
                ])
            
            tab2 = Table(displayName="TableRecap", ref=f"A1:{get_column_letter(ws_recap.max_column)}{ws_recap.max_row}")
            style2 = TableStyleInfo(name="TableStyleMedium4", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab2.tableStyleInfo = style2
            ws_recap.add_table(tab2)

            for ws in [ws_pieces, ws_recap]:
                for col in ws.columns:
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

            # Nom dynamique fichier
            try:
                num_devis = uc.get_project_user_attribute(1).strip().replace(" ", "_")
                client = uc.get_project_number().strip().replace(" ", "_")
                nom_fichier = f"{num_devis}-{client}-prix_configure.xlsx"
            except Exception:
                nom_fichier = "prix_configure.xlsx"

            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            output_path = os.path.join(desktop, nom_fichier)
            wb.save(output_path)
            
            # Ouvrir le fichier
            subprocess.Popen(['start', '', output_path], shell=True)
            self.log(f"📁 Fichier sauvé : {nom_fichier}")

        except Exception as e:
            raise Exception(f"Erreur Excel : {e}")

def main():
    """Fonction principale"""
    app = ConfigurateurPrixInterface()
    app.root.mainloop()

if __name__ == "__main__":
    main()